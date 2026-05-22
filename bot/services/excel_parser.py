"""Excel file parser — extracts lottery numbers from uploaded spreadsheets."""

import io
import re
import logging
from typing import NamedTuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_LOTTERY_PATTERN = re.compile(r"^\d{6}$")


class ParseResult(NamedTuple):
    """Result of parsing an Excel file."""
    numbers: list[str]
    invalid_count: int
    total_rows: int


def parse_excel(file_bytes: bytes) -> ParseResult:
    """
    Parse an Excel file and extract 6-digit lottery numbers.

    Strategy:
    1. Open the first worksheet.
    2. Scan each column to find the one containing the most 6-digit numbers.
    3. Extract all valid 6-digit numbers from that column.
    4. Deduplicate while preserving order.

    Args:
        file_bytes: Raw bytes of the .xlsx / .xls file.

    Returns:
        ParseResult with valid numbers, invalid count, and total rows scanned.
    """
    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
        ws = wb.active

        if ws is None:
            return ParseResult(numbers=[], invalid_count=0, total_rows=0)

        # Read all data
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return ParseResult(numbers=[], invalid_count=0, total_rows=0)

        # Determine the best column (most 6-digit matches)
        num_cols = max(len(row) for row in rows)
        col_scores = [0] * num_cols

        for row in rows:
            for col_idx in range(min(len(row), num_cols)):
                cell = row[col_idx]
                if cell is not None:
                    val = str(cell).strip()
                    # Zero-pad numbers shorter than 6 digits
                    if val.isdigit():
                        val = val.zfill(6)
                    if _LOTTERY_PATTERN.match(val):
                        col_scores[col_idx] += 1

        if max(col_scores, default=0) == 0:
            # No column has valid lottery numbers — try to extract from all cells
            return _extract_from_all_cells(rows)

        best_col = col_scores.index(max(col_scores))

        # Extract numbers from the best column
        numbers = []
        seen = set()
        invalid_count = 0
        total_rows = 0

        for row in rows:
            if best_col >= len(row):
                continue
            cell = row[best_col]
            if cell is None:
                continue

            total_rows += 1
            val = str(cell).strip()

            # Skip header-like values
            if not val.replace(" ", "").isdigit() and not _LOTTERY_PATTERN.match(val):
                # Probably a header
                continue

            # Zero-pad
            if val.isdigit():
                val = val.zfill(6)

            if _LOTTERY_PATTERN.match(val):
                if val not in seen:
                    seen.add(val)
                    numbers.append(val)
            else:
                invalid_count += 1

        wb.close()
        logger.info("Excel parsed: %d valid, %d invalid from %d rows.", len(numbers), invalid_count, total_rows)
        return ParseResult(numbers=numbers, invalid_count=invalid_count, total_rows=total_rows)

    except Exception as e:
        logger.error("Failed to parse Excel file: %s", e)
        return ParseResult(numbers=[], invalid_count=0, total_rows=0)


def _extract_from_all_cells(rows: list) -> ParseResult:
    """Fallback: scan ALL cells for 6-digit numbers."""
    numbers = []
    seen = set()
    invalid_count = 0
    total_cells = 0

    for row in rows:
        for cell in row:
            if cell is None:
                continue
            total_cells += 1
            val = str(cell).strip()
            if val.isdigit():
                val = val.zfill(6)
            if _LOTTERY_PATTERN.match(val):
                if val not in seen:
                    seen.add(val)
                    numbers.append(val)
            elif val.isdigit():
                invalid_count += 1

    return ParseResult(numbers=numbers, invalid_count=invalid_count, total_rows=total_cells)
